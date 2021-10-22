import pandas as pd
import datetime
from openpyxl import load_workbook
import os

THIS_FOLDER = os.path.dirname(os.path.abspath(__file__))
dat_file = os.path.join(THIS_FOLDER, 'result.DAT')


class Converter:

    def __init__(self, file):
        self.file = file
        # LOAD EXCEL FILE
        self.wb = load_workbook(file, data_only=True)
        self.ws = self.wb.worksheets[0]
        self.error_msgs = '<strong>Errors:</strong><br>'
        self.has_error = False

        # GET FILER DATA
        try:
            self.RDO_CODE = f"{self.ws['B1'].value:0>3d}"
        except Exception:
            self.has_error = True
            self.error_msgs += 'Please fill-up RDO Code<br>'

        try:
            self.TIN = f"{self.ws['B7'].value.replace('-','')}"
        except AttributeError:
            try:
                self.TIN = f"{self.ws['B7'].value:0>9d}"
            except Exception:
                self.has_error = True
                self.error_msgs += 'Please fill-up Filer TIN<br>'

        self.NAME = self.ws['B8'].value
        self.TRADE = self.ws['B12'].value
        self.ADDRESS1 = self.ws['B13'].value
        self.ADDRESS2 = self.ws['B14'].value

    def slsp(self):
        try:
            self.ws['B2'].number_format = "m/dd/yyyy"
            self.PERIOD = datetime.datetime.strftime(self.ws['B2'].value, "%m/%d/%Y")
        except (ValueError, TypeError):
            self.has_error = True
            self.error_msgs += 'Please fill-up PERIOD<br>'

        self.CALENDAR = self.ws['B3'].value
        # GET FILER TOTALS
        EXEMPT = f"{self.ws['B18'].value:.2f}"
        ZERO_RATED = f"{self.ws['B19'].value:.2f}"
        SERVICES = f"{self.ws['B20'].value:.2f}"
        CAPITAL_GOODS = f"{self.ws['B21'].value:.2f}"
        GOODS = f"{self.ws['B22'].value:.2f}"
        INPUT_VAT = f"{self.ws['B23'].value:.2f}"
        CREDITABLE = f"{self.ws['B24'].value:.2f}"
        NON_CREDITABLE = f"{self.ws['B25'].value:.2f}"
        NO_OF_RECORDS = f"{self.ws['B26'].value:.2f}"

        # READ AND FIX DATA
        df = pd.read_excel(self.file, 'DATA').fillna(0)
        df['ADDRESS 1'] = df['ADDRESS 1'].replace(0, '-')
        df['ADDRESS 2'] = df['ADDRESS 2'].replace(0, '-')
        df['LAST NAME'] = df['LAST NAME'].replace(0, '')
        df['FIRST NAME'] = df['FIRST NAME'].replace(0, '')
        df['MIDDLE NAME'] = df['MIDDLE NAME'].replace(0, '')

        # CONVERT AND FORMAT DATA
        def parse(line):
            try:
                line[0] = f"{line[0].replace('-','')}"
            except AttributeError:
                line[0] = f"{line[0]:0>9d}"
                print(line[0])
                print(type(line[0]))

            line[7] = f'{line[7]:.2f}'
            line[8] = f'{line[8]:.2f}'
            line[9] = f'{line[9]:.2f}'
            line[10] = f'{line[10]:.2f}'
            line[11] = f'{line[11]:.2f}'
            line[12] = f'{line[12]:.2f}'

            for i in range(len(line)):
                try:
                    line[i] = line[i].upper()
                except AttributeError:
                    pass

            return f'D,P,"{line[0]}","{line[1]}",,,,"{line[5]}","{line[6]}",{line[7]},{line[8]},{line[9]},{line[10]},{line[11]},{line[12]},{self.TIN},{self.PERIOD}\n'

        # RETURN RESULT
        if self.has_error:
            return self.error_msgs
        else:
            with open(dat_file, 'w') as dat:
                dat.write(f'H,P,"{self.TIN}","{self.NAME}","","","","{self.TRADE}","{self.ADDRESS1}","{self.ADDRESS2}",{EXEMPT},{ZERO_RATED},{SERVICES},{CAPITAL_GOODS},{GOODS},{INPUT_VAT},{CREDITABLE},{NON_CREDITABLE},{self.RDO_CODE},{self.PERIOD},{self.CALENDAR}\n')
                dat.writelines([parse(line) for line in df.values])

            dest_path = f'{self.TIN}P{self.PERIOD[:2]}{self.PERIOD[6:10]}.DAT'
            return dest_path

    def qap(self):
        BRANCH_CODE = f"{self.ws['D7'].value:0>4d}"
        TOTAL_PAYMENTS = f"{self.ws['B15'].value:.2f}"
        TOTAL_WITHHELD = f"{self.ws['B16'].value:.2f}"
        self.PERIOD = f"{self.ws['B3'].value:0>2d}/{self.ws['B2'].value}"

        df = pd.read_excel(self.file, 'DATA').fillna(0)
        df['PAYEE REGISTERED NAME'] = df['PAYEE REGISTERED NAME'].replace(0, '')

        def parse(line, idx):
            try:
                line[0] = f"{line[0].replace('-','')}"
            except AttributeError:
                line[0] = f"{line[0]:0>9d}"
                print(line[0])
                print(type(line[0]))

            line[1] = f'{line[1]:0>4d}'
            line[4] = f'{line[4]:.2f}'
            line[5] = f'{line[5]:.2f}'
            line[6] = f'{line[6]:.2f}'

            for i in range(len(line)):
                try:
                    line[i] = line[i].upper()
                except AttributeError:
                    pass

            return f'D1,1601EQ,{idx+1},{line[0]},{line[1]},"{line[2]}",,,,{self.PERIOD},{line[3]},{line[4]},{line[5]},{line[6]}\n'

        # RETURN RESULT
        if self.has_error:
            return self.error_msgs
        else:
            with open(dat_file, 'w') as dat:
                dat.write(f'HQAP,H1601EQ,{self.TIN},{BRANCH_CODE},"{self.NAME}",{self.PERIOD},{self.RDO_CODE}\n')

                data = []
                for i in range(len(df.values)):
                    data.append(parse(df.values[i], i))
                dat.writelines(data)

                dat.write(f'C1,1601EQ,{self.TIN},{BRANCH_CODE},{self.PERIOD},{TOTAL_PAYMENTS},{TOTAL_WITHHELD}\n')

            dest_path = f'{self.TIN}{BRANCH_CODE}{self.PERIOD[:2]}{self.PERIOD[3:7]}1601EQ.DAT'
            return dest_path
